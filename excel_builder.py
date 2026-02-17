"""Build organized Excel workbooks from extracted SEC filing data."""

import os
import re
import tempfile
from datetime import datetime

from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.utils import get_column_letter


# Default colors (overridden by brand_colors)
DEFAULT_PRIMARY = "4472C4"
DEFAULT_ACCENT = "D9E1F2"

THIN_BORDER = Border(
    bottom=Side(style="thin", color="CCCCCC"),
)
SUBTOTAL_BORDER = Border(
    top=Side(style="thin", color="000000"),
    bottom=Side(style="thin", color="000000"),
)
TOTAL_BORDER = Border(
    top=Side(style="thin", color="000000"),
    bottom=Side(style="double", color="000000"),
)

# Accounting format
ACCT_FMT = '#,##0_);(#,##0)'
ACCT_FMT_DEC = '#,##0.00_);(#,##0.00)'
PCT_FMT = '0.00%'


# ─── Statement Structures for Formula-Based Output ─────────────────────
#
# Each item is one of:
#   "data"    — writes the XBRL value from the parsed data
#   "formula" — writes an Excel formula referencing other rows
#   "section" — writes a bold section header (e.g. "ASSETS")
#   "spacer"  — empty row for visual separation
#
# Formula items:
#   "plus"     — list of item IDs whose cell values are added
#   "minus"    — list of item IDs whose cell values are subtracted
#   "fallback" — labels to look up in XBRL data if no formula components exist
#   "total"    — if True, use double-underline border (major total)
#
# Data items:
#   "labels"  — possible XBRL labels (tries each until one is found)
#   "negate"  — if True, flip sign for display (e.g. payments shown as negative)

INCOME_STRUCTURE = [
    {"id": "revenue", "labels": ["Revenue"], "type": "data"},
    {"id": "cogs", "labels": ["Cost of Revenue", "Cost of Goods Sold"], "type": "data"},
    {"id": "gross_profit", "label": "Gross Profit", "type": "formula",
     "plus": ["revenue"], "minus": ["cogs"],
     "fallback": ["Gross Profit"]},

    {"type": "spacer"},

    {"id": "rd", "labels": ["Research & Development"], "type": "data"},
    {"id": "sga", "labels": ["Selling, General & Administrative", "Selling & Marketing",
                              "General & Administrative"], "type": "data"},
    {"id": "opex", "label": "Total Operating Expenses", "type": "formula",
     "plus": ["rd", "sga"], "minus": [],
     "fallback": ["Total Operating Expenses"]},

    {"type": "spacer"},

    {"id": "op_income", "label": "Operating Income (Loss)", "type": "formula",
     "plus": ["gross_profit"], "minus": ["opex"],
     "fallback": ["Operating Income (Loss)"]},

    {"type": "spacer"},

    {"id": "int_exp", "labels": ["Interest Expense"], "type": "data"},
    {"id": "int_inc", "labels": ["Interest Income"], "type": "data"},
    {"id": "int_net", "labels": ["Net Interest Income (Expense)"], "type": "data"},
    {"id": "other_nonop", "labels": ["Other Non-Operating Income (Expense)"], "type": "data"},
    {"id": "pretax", "label": "Income Before Tax", "type": "formula",
     "plus": ["op_income", "int_inc", "int_net", "other_nonop"],
     "minus": ["int_exp"],
     "fallback": ["Income Before Tax"]},

    {"type": "spacer"},

    {"id": "tax", "labels": ["Income Tax Expense (Benefit)"], "type": "data"},
    {"id": "net_income", "label": "Net Income (Loss)", "type": "formula",
     "plus": ["pretax"], "minus": ["tax"],
     "fallback": ["Net Income (Loss)"], "total": True},

    {"type": "spacer"},

    {"id": "eps_basic", "labels": ["EPS (Basic)"], "type": "data"},
    {"id": "eps_diluted", "labels": ["EPS (Diluted)"], "type": "data"},
    {"id": "shares_bd", "labels": ["Weighted Avg Shares (Basic & Diluted)"], "type": "data"},
    {"id": "shares_b", "labels": ["Weighted Avg Shares (Basic)"], "type": "data"},
    {"id": "shares_d", "labels": ["Weighted Avg Shares (Diluted)"], "type": "data"},
]

BALANCE_STRUCTURE = [
    {"type": "section", "label": "ASSETS"},

    {"id": "cash", "labels": ["Cash & Cash Equivalents"], "type": "data"},
    {"id": "st_inv", "labels": ["Short-Term Investments"], "type": "data"},
    {"id": "ar", "labels": ["Accounts Receivable"], "type": "data"},
    {"id": "inv", "labels": ["Inventory"], "type": "data"},
    {"id": "prepaid", "labels": ["Prepaid Expenses & Other Current Assets"], "type": "data"},
    {"id": "ca", "label": "Total Current Assets", "type": "formula",
     "plus": ["cash", "st_inv", "ar", "inv", "prepaid"], "minus": [],
     "fallback": ["Total Current Assets"]},

    {"type": "spacer"},

    {"id": "ppe", "labels": ["Property, Plant & Equipment (Net)"], "type": "data"},
    {"id": "gw", "labels": ["Goodwill"], "type": "data"},
    {"id": "intang", "labels": ["Intangible Assets (Net)"], "type": "data"},
    {"id": "other_nca", "labels": ["Other Non-Current Assets"], "type": "data"},
    {"id": "total_assets", "label": "Total Assets", "type": "formula",
     "plus": ["ca", "ppe", "gw", "intang", "other_nca"], "minus": [],
     "fallback": ["Total Assets"], "total": True},

    {"type": "spacer"},
    {"type": "section", "label": "LIABILITIES"},

    {"id": "ap", "labels": ["Accounts Payable"], "type": "data"},
    {"id": "accrued", "labels": ["Accrued Liabilities"], "type": "data"},
    {"id": "cur_ltd", "labels": ["Current Portion of Long-Term Debt"], "type": "data"},
    {"id": "st_borr", "labels": ["Short-Term Borrowings"], "type": "data"},
    {"id": "cl", "label": "Total Current Liabilities", "type": "formula",
     "plus": ["ap", "accrued", "cur_ltd", "st_borr"], "minus": [],
     "fallback": ["Total Current Liabilities"]},

    {"type": "spacer"},

    {"id": "lt_debt", "labels": ["Long-Term Debt"], "type": "data"},
    {"id": "lease", "labels": ["Operating Lease Liabilities (Non-Current)"], "type": "data"},
    {"id": "other_ncl", "labels": ["Other Non-Current Liabilities"], "type": "data"},
    {"id": "total_liab", "label": "Total Liabilities", "type": "formula",
     "plus": ["cl", "lt_debt", "lease", "other_ncl"], "minus": [],
     "fallback": ["Total Liabilities"]},

    {"type": "spacer"},
    {"type": "section", "label": "STOCKHOLDERS' EQUITY"},

    {"id": "cs", "labels": ["Common Stock"], "type": "data"},
    {"id": "apic", "labels": ["Additional Paid-In Capital"], "type": "data"},
    {"id": "re", "labels": ["Retained Earnings (Accumulated Deficit)"], "type": "data"},
    {"id": "aoci", "labels": ["Accumulated Other Comprehensive Income (Loss)"], "type": "data"},
    {"id": "treasury", "labels": ["Treasury Stock"], "type": "data", "negate": True},
    {"id": "total_eq", "label": "Total Stockholders' Equity", "type": "formula",
     "plus": ["cs", "apic", "re", "aoci", "treasury"], "minus": [],
     "fallback": ["Total Stockholders' Equity"]},

    {"type": "spacer"},

    {"id": "total_le", "label": "Total Liabilities & Stockholders' Equity", "type": "formula",
     "plus": ["total_liab", "total_eq"], "minus": [],
     "fallback": ["Total Liabilities & Stockholders' Equity"], "total": True},
]

CASHFLOW_STRUCTURE = [
    {"type": "section", "label": "OPERATING ACTIVITIES"},

    {"id": "cf_ni", "labels": ["Net Income"], "type": "data"},
    {"id": "da", "labels": ["Depreciation & Amortization"], "type": "data"},
    {"id": "sbc", "labels": ["Stock-Based Compensation"], "type": "data"},
    {"id": "def_tax", "labels": ["Deferred Income Taxes"], "type": "data"},
    {"id": "chg_ar", "labels": ["Change in Accounts Receivable"], "type": "data"},
    {"id": "chg_inv", "labels": ["Change in Inventories"], "type": "data"},
    {"id": "chg_ap", "labels": ["Change in Accounts Payable"], "type": "data"},
    {"id": "chg_acc", "labels": ["Change in Accrued Liabilities"], "type": "data"},
    {"id": "cfo", "label": "Net Cash from Operating Activities", "type": "formula",
     "plus": ["cf_ni", "da", "sbc", "def_tax", "chg_ar", "chg_inv", "chg_ap", "chg_acc"],
     "minus": [],
     "fallback": ["Net Cash from Operating Activities"]},

    {"type": "spacer"},
    {"type": "section", "label": "INVESTING ACTIVITIES"},

    {"id": "capex", "labels": ["Capital Expenditures"], "type": "data", "negate": True},
    {"id": "acq", "labels": ["Acquisitions (Net of Cash)"], "type": "data", "negate": True},
    {"id": "buy_inv", "labels": ["Purchases of Investments"], "type": "data", "negate": True},
    {"id": "sell_inv", "labels": ["Proceeds from Sale of Investments"], "type": "data"},
    {"id": "mat_inv", "labels": ["Proceeds from Maturities of Investments"], "type": "data"},
    {"id": "cfi", "label": "Net Cash from Investing Activities", "type": "formula",
     "plus": ["capex", "acq", "buy_inv", "sell_inv", "mat_inv"],
     "minus": [],
     "fallback": ["Net Cash from Investing Activities"]},

    {"type": "spacer"},
    {"type": "section", "label": "FINANCING ACTIVITIES"},

    {"id": "debt_proc", "labels": ["Proceeds from Long-Term Debt"], "type": "data"},
    {"id": "debt_repay", "labels": ["Repayments of Long-Term Debt"], "type": "data", "negate": True},
    {"id": "divs", "labels": ["Dividends Paid", "Dividends Paid (Common Stock)"],
     "type": "data", "negate": True},
    {"id": "buybacks", "labels": ["Share Repurchases"], "type": "data", "negate": True},
    {"id": "stock_iss", "labels": ["Proceeds from Stock Issuance"], "type": "data"},
    {"id": "cff", "label": "Net Cash from Financing Activities", "type": "formula",
     "plus": ["debt_proc", "debt_repay", "divs", "buybacks", "stock_iss"],
     "minus": [],
     "fallback": ["Net Cash from Financing Activities"]},

    {"type": "spacer"},

    {"id": "net_cash", "label": "Net Change in Cash", "type": "formula",
     "plus": ["cfo", "cfi", "cff"], "minus": [],
     "fallback": ["Net Change in Cash"], "total": True},
]

STATEMENT_STRUCTURES = {
    "Income Statement": INCOME_STRUCTURE,
    "Balance Sheet": BALANCE_STRUCTURE,
    "Cash Flow": CASHFLOW_STRUCTURE,
}


# ─── Helpers ────────────────────────────────────────────────────────────

def _hex_to_openpyxl(hex_color):
    """Convert #RRGGBB to RRGGBB for openpyxl."""
    return hex_color.lstrip("#").upper()


def _make_styles(primary_hex, accent_hex):
    """Create style objects from brand colors."""
    primary = _hex_to_openpyxl(primary_hex)
    accent = _hex_to_openpyxl(accent_hex)

    return {
        "title_fill": PatternFill(start_color=primary, end_color=primary, fill_type="solid"),
        "title_font": Font(bold=True, size=13, color="FFFFFF"),
        "header_fill": PatternFill(start_color=accent, end_color=accent, fill_type="solid"),
        "header_font": Font(bold=True, size=11),
        "title_font_plain": Font(bold=True, size=13),
        "subtitle_font": Font(bold=True, size=11, color="555555"),
    }


def _safe_sheet_name(name, max_len=31):
    """Make a string safe for use as an Excel sheet name."""
    name = re.sub(r'[\\/*?\[\]:]', '', name)
    if len(name) > max_len:
        name = name[:max_len]
    return name.strip()


def _unique_sheet_name(name, used_names):
    """Ensure a sheet name is unique by appending a counter if needed."""
    base = _safe_sheet_name(name)
    if not base:
        base = "Table"
    result = base
    counter = 2
    while result in used_names:
        suffix = f" ({counter})"
        result = _safe_sheet_name(base[:31 - len(suffix)] + suffix)
        counter += 1
    used_names.add(result)
    return result


def _is_year_like(val):
    """Check if a value looks like a year (1900-2099) and should NOT get comma formatting."""
    if isinstance(val, int) and 1900 <= val <= 2099:
        return True
    if isinstance(val, float) and val == int(val) and 1900 <= int(val) <= 2099:
        return True
    return False


def _try_parse_number(text):
    """Try to convert text to a number for Excel."""
    if not text or text in ("\u2014", "\u2013", "-", ""):
        return None, False

    cleaned = text.replace("$", "").replace(",", "").replace(" ", "").strip()

    # Handle parentheses as negatives: (123) -> -123
    if cleaned.startswith("(") and cleaned.endswith(")"):
        cleaned = "-" + cleaned[1:-1]

    is_pct = cleaned.endswith("%")
    if is_pct:
        cleaned = cleaned[:-1]

    try:
        val = float(cleaned)
        if is_pct:
            val = val / 100.0
            return val, True
        if val == int(val):
            return int(val), False
        return val, False
    except (ValueError, OverflowError):
        return None, False


def _format_number_cell(cell, val, is_pct=False):
    """Apply accounting formatting to a number cell."""
    cell.value = val
    cell.alignment = Alignment(horizontal="right")

    if is_pct:
        cell.number_format = PCT_FMT
    elif _is_year_like(val):
        cell.number_format = '0'
    elif isinstance(val, float):
        cell.number_format = ACCT_FMT_DEC
    elif isinstance(val, int):
        cell.number_format = ACCT_FMT


def _auto_width(ws, min_width=12, max_width=45):
    """Auto-fit column widths based on content."""
    for col_cells in ws.columns:
        max_length = min_width
        col_letter = get_column_letter(col_cells[0].column)
        for cell in col_cells:
            if cell.value:
                length = len(str(cell.value)) + 3
                max_length = max(max_length, min(length, max_width))
        ws.column_dimensions[col_letter].width = max_length


def _write_title_bar(ws, row, title, num_cols, styles):
    """Write a colored title bar spanning multiple columns."""
    ws.cell(row=row, column=1, value=title).font = styles["title_font"]
    ws.cell(row=row, column=1).fill = styles["title_fill"]
    for col in range(2, num_cols + 1):
        ws.cell(row=row, column=col).fill = styles["title_fill"]


def _find_data_for_item(statement_data, labels):
    """Find period->value dict for an item by checking multiple possible labels.

    Returns (values_dict, matched_label) or (None, None).
    """
    for label in labels:
        if label in statement_data:
            return statement_data[label], label
    return None, None


def _build_formula(col_letter, plus_rows, minus_rows):
    """Build an Excel formula string from plus/minus row references.

    Example: _build_formula("B", [3, 5], [4]) -> "=B3+B5-B4"
    """
    if not plus_rows and not minus_rows:
        return None

    parts = []
    for r in plus_rows:
        parts.append(f"+{col_letter}{r}")
    for r in minus_rows:
        parts.append(f"-{col_letter}{r}")

    formula = "=" + "".join(parts)
    # Clean up leading +
    if formula.startswith("=+"):
        formula = "=" + formula[2:]
    return formula


# ─── Formula-Based Statement Writer ────────────────────────────────────

def _write_formula_statement(ws, statement_name, statement_data, periods, start_row, styles):
    """Write a financial statement with real Excel formulas. Returns next available row.

    Uses the structured layout from STATEMENT_STRUCTURES. Data items get XBRL values,
    formula items get Excel formulas (e.g. =B3-B4). Falls back to XBRL values when
    formula components are unavailable.
    """
    structure = STATEMENT_STRUCTURES.get(statement_name)
    if not structure or not statement_data:
        return start_row

    row = start_row
    num_cols = len(periods) + 1

    # Title bar
    _write_title_bar(ws, row, statement_name, num_cols, styles)
    row += 1

    # Period headers
    ws.cell(row=row, column=1, value="Line Item").font = styles["header_font"]
    ws.cell(row=row, column=1).fill = styles["header_fill"]
    for i, period in enumerate(periods):
        cell = ws.cell(row=row, column=i + 2, value=period)
        cell.font = styles["header_font"]
        cell.fill = styles["header_fill"]
        cell.alignment = Alignment(horizontal="center")
    row += 1

    # Track which Excel row each item ID is written to
    row_map = {}

    for item in structure:
        item_type = item.get("type")

        # ── Spacer ──
        if item_type == "spacer":
            row += 1
            continue

        # ── Section header ──
        if item_type == "section":
            ws.cell(row=row, column=1, value=item["label"]).font = Font(
                bold=True, size=11, color="555555"
            )
            row += 1
            continue

        item_id = item.get("id")

        # ── Data item ──
        if item_type == "data":
            values, matched_label = _find_data_for_item(
                statement_data, item.get("labels", [])
            )
            if not values:
                continue  # Skip items with no XBRL data

            negate = item.get("negate", False)
            display_label = matched_label or item.get("labels", [""])[0]

            c = ws.cell(row=row, column=1, value=display_label)
            c.border = THIN_BORDER
            c.alignment = Alignment(indent=1)

            has_any_value = False
            for i, period in enumerate(periods):
                if period in values:
                    val = values[period]
                    if isinstance(val, (int, float)):
                        if negate:
                            val = -val
                        _format_number_cell(ws.cell(row=row, column=i + 2), val)
                        has_any_value = True

            if has_any_value:
                row_map[item_id] = row
            row += 1

        # ── Formula item ──
        elif item_type == "formula":
            plus_rows = [row_map[ref] for ref in item.get("plus", []) if ref in row_map]
            minus_rows = [row_map[ref] for ref in item.get("minus", []) if ref in row_map]
            has_formula = bool(plus_rows or minus_rows)

            is_total = item.get("total", False)
            border = TOTAL_BORDER if is_total else SUBTOTAL_BORDER

            if has_formula:
                label = item.get("label", "")
                c = ws.cell(row=row, column=1, value=label)
                c.font = Font(bold=True)
                c.border = border

                for i, period in enumerate(periods):
                    col_letter = get_column_letter(i + 2)
                    formula = _build_formula(col_letter, plus_rows, minus_rows)
                    if formula:
                        cell = ws.cell(row=row, column=i + 2)
                        cell.value = formula
                        cell.number_format = ACCT_FMT
                        cell.alignment = Alignment(horizontal="right")
                        cell.font = Font(bold=True)
                        cell.border = border

                row_map[item_id] = row
                row += 1

            else:
                # No formula components available — fall back to XBRL value
                fallback_labels = item.get("fallback", [])
                values, matched_label = _find_data_for_item(statement_data, fallback_labels)
                if values:
                    label = item.get("label", "")
                    c = ws.cell(row=row, column=1, value=label)
                    c.font = Font(bold=True)
                    c.border = border

                    has_any_value = False
                    for i, period in enumerate(periods):
                        if period in values:
                            val = values[period]
                            if isinstance(val, (int, float)):
                                cell = ws.cell(row=row, column=i + 2)
                                _format_number_cell(cell, val)
                                cell.font = Font(bold=True)
                                cell.border = border
                                has_any_value = True

                    if has_any_value:
                        row_map[item_id] = row
                    row += 1

    row += 1
    return row


# ─── HTML Table Writer (unchanged) ─────────────────────────────────────

def _write_html_table(ws, table_dict, start_row, styles, title_override=None, filing_source=None):
    """Write one HTML-extracted table with matching formatting. Returns next available row."""
    row = start_row
    title = title_override or table_dict.get("title") or "Table"
    headers = table_dict.get("headers", [])
    data_rows = table_dict.get("rows", [])

    # Determine column count
    max_cols = 1
    for hr in headers:
        max_cols = max(max_cols, len(hr))
    for dr in data_rows:
        max_cols = max(max_cols, len(dr))

    # Title bar
    _write_title_bar(ws, row, title, max_cols, styles)
    row += 1

    # Source subtitle if provided
    if filing_source:
        ws.cell(row=row, column=1, value=f"Source: {filing_source}").font = styles["subtitle_font"]
        row += 1

    # Headers
    for header_row in headers:
        for col_idx, val in enumerate(header_row):
            cell = ws.cell(row=row, column=col_idx + 1, value=val)
            cell.font = styles["header_font"]
            cell.fill = styles["header_fill"]
            cell.alignment = Alignment(horizontal="center")
        row += 1

    # Data rows
    for data_row in data_rows:
        for col_idx, val in enumerate(data_row):
            num, is_pct = _try_parse_number(val)
            if num is not None:
                _format_number_cell(ws.cell(row=row, column=col_idx + 1), num, is_pct)
            else:
                ws.cell(row=row, column=col_idx + 1, value=val)
                if col_idx == 0:
                    ws.cell(row=row, column=col_idx + 1).border = THIN_BORDER
        row += 1

    row += 1
    return row


# ─── Workbook Builders ─────────────────────────────────────────────────

def build_workbook(company_name, ticker, xbrl_data, selected_tables, selected_filings,
                   single_sheet=False, brand_colors=None):
    """Build and save an Excel workbook.

    Args:
        company_name: Company display name.
        ticker: Company ticker symbol.
        xbrl_data: Output from xbrl_parser.extract_financials().
        selected_tables: List of {table, filing_type, filing_date} dicts.
        selected_filings: List of filing dicts {type, date, accession, ...}.
        single_sheet: If True, put everything on one sheet instead of separate tabs.
        brand_colors: Optional dict with 'primary' and 'accent' hex colors.

    Returns:
        (filepath, filename) tuple.
    """
    # Build styles from brand colors
    primary = (brand_colors or {}).get("primary", "#" + DEFAULT_PRIMARY)
    accent = (brand_colors or {}).get("accent", "#" + DEFAULT_ACCENT)
    styles = _make_styles(primary, accent)

    wb = Workbook()
    used_sheet_names = set()
    periods = xbrl_data.get("periods", [])

    if single_sheet:
        _build_single_sheet(wb, company_name, ticker, xbrl_data, selected_tables, selected_filings, periods, styles)
    else:
        _build_multi_sheet(wb, company_name, ticker, xbrl_data, selected_tables, selected_filings, periods, used_sheet_names, styles)

    # Save
    safe_ticker = re.sub(r'[^A-Za-z0-9]', '', ticker or company_name[:10])
    filename = f"{safe_ticker}_SEC_Filings.xlsx"
    output_dir = tempfile.mkdtemp()
    filepath = os.path.join(output_dir, filename)
    wb.save(filepath)

    return filepath, filename


def _build_single_sheet(wb, company_name, ticker, xbrl_data, selected_tables, selected_filings, periods, styles):
    """Put all data on a single sheet, one section after another."""
    ws = wb.active
    ws.title = "All Data"

    row = 1
    ws.cell(row=row, column=1, value=f"{company_name} ({ticker})").font = styles["title_font_plain"]
    row += 1
    ws.cell(row=row, column=1, value=f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M')}").font = styles["subtitle_font"]
    row += 2

    # XBRL statements — now with formulas
    for statement_name in ("Income Statement", "Balance Sheet", "Cash Flow"):
        statement_data = xbrl_data.get(statement_name, {})
        if statement_data:
            row = _write_formula_statement(ws, statement_name, statement_data, periods, start_row=row, styles=styles)
            row += 1

    # Selected HTML tables
    for entry in selected_tables:
        table = entry["table"]
        filing_type = entry.get("filing_type", "")
        filing_date = entry.get("filing_date", "")
        source = f"{filing_type} ({filing_date})"
        row = _write_html_table(ws, table, start_row=row, styles=styles, filing_source=source)
        row += 1

    ws.freeze_panes = "B1"
    _auto_width(ws)


def _build_multi_sheet(wb, company_name, ticker, xbrl_data, selected_tables, selected_filings, periods, used_sheet_names, styles):
    """Separate tabs: Index + core financials + one sheet per selected table."""

    # --- Index Sheet ---
    ws_index = wb.active
    ws_index.title = "Index"
    used_sheet_names.add("Index")

    ws_index.cell(row=1, column=1, value=f"{company_name} ({ticker})").font = styles["title_font_plain"]
    ws_index.cell(row=2, column=1, value=f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M')}").font = styles["subtitle_font"]

    ws_index.cell(row=4, column=1, value="Filings Included:").font = styles["header_font"]
    row = 5
    ws_index.cell(row=row, column=1, value="Type").font = styles["header_font"]
    ws_index.cell(row=row, column=1).fill = styles["header_fill"]
    ws_index.cell(row=row, column=2, value="Date").font = styles["header_font"]
    ws_index.cell(row=row, column=2).fill = styles["header_fill"]
    row += 1
    for filing in selected_filings:
        ws_index.cell(row=row, column=1, value=filing.get("type", ""))
        ws_index.cell(row=row, column=2, value=filing.get("date", ""))
        row += 1

    if selected_tables:
        row += 1
        ws_index.cell(row=row, column=1, value="Additional Tables Included:").font = styles["header_font"]
        row += 1
        ws_index.cell(row=row, column=1, value="Sheet Name").font = styles["header_font"]
        ws_index.cell(row=row, column=1).fill = styles["header_fill"]
        ws_index.cell(row=row, column=2, value="Source").font = styles["header_font"]
        ws_index.cell(row=row, column=2).fill = styles["header_fill"]
        row += 1

    _auto_width(ws_index)

    # --- Core Financial Statement Sheets (with formulas) ---
    for statement_name in ("Income Statement", "Balance Sheet", "Cash Flow"):
        statement_data = xbrl_data.get(statement_name, {})
        if not statement_data:
            continue

        sheet_name = _unique_sheet_name(statement_name, used_sheet_names)
        ws = wb.create_sheet(title=sheet_name)
        ws.freeze_panes = "B3"
        _write_formula_statement(ws, statement_name, statement_data, periods, start_row=1, styles=styles)
        _auto_width(ws)

    # --- Individual Table Sheets ---
    index_table_row = row
    for entry in selected_tables:
        table = entry["table"]
        filing_type = entry.get("filing_type", "")
        filing_date = entry.get("filing_date", "")
        title = table.get("title") or "Table"
        source = f"{filing_type} ({filing_date})"

        sheet_name = _unique_sheet_name(title, used_sheet_names)
        ws = wb.create_sheet(title=sheet_name)
        ws.freeze_panes = "A3"

        _write_html_table(ws, table, start_row=1, styles=styles, filing_source=source)
        _auto_width(ws)

        # Add to index
        ws_index.cell(row=index_table_row, column=1, value=sheet_name)
        ws_index.cell(row=index_table_row, column=2, value=source)
        index_table_row += 1

    _auto_width(ws_index)
